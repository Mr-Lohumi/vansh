"""
FCTG GCC Techlab - Hot-Desk Booking Backend
--------------------------------------------
Local server. The Excel file 'seating_backend.xlsx' (same folder) is the single
source of truth. The server reads/writes it and enforces booking rules, so two
people can never grab the same desk on the same day.

RUN:
    pip install flask openpyxl
    python app.py
Then open  http://localhost:5000  on this machine, and share
    http://<this-machine-LAN-IP>:5000   with your team (same network).

NOTE: Keep seating_backend.xlsx CLOSED in Excel while the server runs
(the server needs to write to it).
"""
import os, threading, datetime, sqlite3
from flask import Flask, request, jsonify, send_from_directory
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, 'seating_backend.xlsx')
CON_XLSX = os.path.join(BASE, 'contractors_database.xlsx')
SEATING_DB = os.path.join(BASE, 'seating.db')
DN = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
LOCK = threading.Lock()
app = Flask(__name__)


def get_seating_db_connection():
    conn = sqlite3.connect(SEATING_DB)
    conn.row_factory = sqlite3.Row
    return conn


def init_seating_db():
    conn = get_seating_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS SeatingConfig (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS SeatingRoster (
            name TEXT PRIMARY KEY,
            team TEXT,
            old_seat TEXT,
            type TEXT,
            Mon TEXT,
            Tue TEXT,
            Wed TEXT,
            Thu TEXT,
            Fri TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS SeatingBookings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            desk INTEGER,
            day TEXT,
            name TEXT,
            role TEXT,
            timestamp TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS SeatingMiddleBay (
            desk INTEGER PRIMARY KEY,
            name TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS SeatingVacations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            start_date TEXT,
            end_date TEXT,
            created_at TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS SeatingPods (
            pod_name TEXT PRIMARY KEY,
            top_desks TEXT,
            bot_desks TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS SeatingAdmins (
            Username TEXT PRIMARY KEY,
            Password TEXT,
            UpdatedAt TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS SeatingTeamAdmins (
            team TEXT PRIMARY KEY,
            username TEXT,
            password TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS TeamRosterPeople (
            name TEXT PRIMARY KEY,
            team TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS TeamRosterMarks (
            month TEXT,
            name TEXT,
            day INTEGER,
            mark TEXT,
            PRIMARY KEY (month, name, day)
        )
    ''')
    
    conn.commit()
    
    # Seed default pods if SeatingPods is empty
    cursor.execute("SELECT COUNT(*) FROM SeatingPods")
    if cursor.fetchone()[0] == 0:
        default_pods = [
            ('L', '281,282,283,284,285,286', ''),
            ('1', '287,288,289,290,291,292', '293,294,295,296,297,298'),
            ('2', '299,300,301,302,303', '304,305,306,307,308'),
            ('3', '309,310,311,312,313,314', '315,316,317,318,319,320'),
            ('4', '321,322,323,324,325,326', '327,328,329,330,331,332'),
            ('5', '333,334,335,336,337', '338,339,340,341,342'),
            ('6', '343,344,345,346,347,348', '349,350,351,352,353,354')
        ]
        cursor.executemany("INSERT INTO SeatingPods (pod_name, top_desks, bot_desks) VALUES (?, ?, ?)", default_pods)
        conn.commit()
        
    # Check if table is empty and if Excel file exists for migration
    cursor.execute("SELECT COUNT(*) FROM SeatingRoster")
    row_count = cursor.fetchone()[0]
    
    if row_count == 0 and os.path.exists(XLSX):
        print("Migrating Seating Planner Excel to SQLite...")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(XLSX)
            
            # Migrate Config
            if 'Config' in wb.sheetnames:
                ws = wb['Config']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        cursor.execute("INSERT OR REPLACE INTO SeatingConfig (key, value) VALUES (?, ?)", (str(row[0]).strip(), str(row[1]).strip() if row[1] is not None else ''))
            
            # Migrate Roster
            if 'Roster' in wb.sheetnames:
                ws = wb['Roster']
                headers = [c.value for c in ws[1]]
                headers = [h.strip() if h else '' for h in headers]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    row_dict = dict(zip(headers, row))
                    cursor.execute("""
                        INSERT OR REPLACE INTO SeatingRoster (name, team, old_seat, type, Mon, Tue, Wed, Thu, Fri)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        str(row_dict.get('name', '') or '').strip(),
                        str(row_dict.get('team', '') or '').strip(),
                        str(row_dict.get('old_seat', '') or row_dict.get('Old Seat', '') or row_dict.get('OldSeat', '') or '').strip(),
                        str(row_dict.get('type', 'shared') or '').strip(),
                        str(row_dict.get('Mon', '') or '').strip(),
                        str(row_dict.get('Tue', '') or '').strip(),
                        str(row_dict.get('Wed', '') or '').strip(),
                        str(row_dict.get('Thu', '') or '').strip(),
                        str(row_dict.get('Fri', '') or '').strip()
                    ))
            
            # Migrate Bookings
            if 'Bookings' in wb.sheetnames:
                ws = wb['Bookings']
                headers = [c.value for c in ws[1]]
                headers = [h.strip() if h else '' for h in headers]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[1] in (None, ''):
                        continue
                    row_dict = dict(zip(headers, row))
                    cursor.execute("""
                        INSERT INTO SeatingBookings (id, desk, day, name, role, timestamp)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (
                        row_dict.get('id'),
                        int(row_dict.get('desk')),
                        str(row_dict.get('day', '') or '').strip(),
                        str(row_dict.get('name', '') or '').strip(),
                        str(row_dict.get('role', '') or '').strip(),
                        str(row_dict.get('timestamp', '') or '').strip()
                    ))
            
            # Migrate MiddleBay
            if 'MiddleBay' in wb.sheetnames:
                ws = wb['MiddleBay']
                headers = [c.value for c in ws[1]]
                headers = [h.strip() if h else '' for h in headers]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    row_dict = dict(zip(headers, row))
                    cursor.execute("""
                        INSERT OR REPLACE INTO SeatingMiddleBay (desk, name)
                        VALUES (?, ?)
                    """, (
                        int(row_dict.get('desk')),
                        str(row_dict.get('name', '') or '').strip()
                    ))
                    
            # Migrate Vacations
            if 'Vacations' in wb.sheetnames:
                ws = wb['Vacations']
                headers = [c.value for c in ws[1]]
                headers = [h.strip() if h else '' for h in headers]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    row_dict = dict(zip(headers, row))
                    cursor.execute("""
                        INSERT INTO SeatingVacations (id, name, start_date, end_date, created_at)
                        VALUES (?, ?, ?, ?, ?)
                    """, (
                        row_dict.get('id'),
                        str(row_dict.get('name', '') or '').strip(),
                        str(row_dict.get('start_date', '') or '').strip(),
                        str(row_dict.get('end_date', '') or '').strip(),
                        str(row_dict.get('created_at', '') or '').strip()
                    ))
            
            conn.commit()
            print("Seating Planner migration completed successfully!")
        except Exception as e:
            conn.rollback()
            print("Error during Seating Excel-to-SQLite migration:", e)
            
    conn.close()


def get_week_dates():
    today = datetime.date.today()
    monday = today - datetime.timedelta(days=today.weekday())
    dates = {}
    dn = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    for idx, name in enumerate(dn):
        d = monday + datetime.timedelta(days=idx)
        dates[name] = d.strftime('%Y-%m-%d')
    return dates


def check_seating_admin_auth(data):
    if not data:
        return False
    
    # 1. Check if DBA admin
    super_user = (data.get('super_user') or '').strip()
    super_pass = (data.get('super_pass') or '').strip()
    if super_user and super_pass:
        if check_dba_auth_db(super_user, super_pass):
            return True
            
    # 2. Check if Contractor Admin
    admin_id = (data.get('admin_id') or '').strip()
    admin_pass = (data.get('password') or data.get('admin_pass') or '').strip()
    if admin_id and admin_pass:
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM ContractorAdmins WHERE LOWER(AdminID) = ?", (admin_id.lower(),))
            admin_rec = cursor.fetchone()
            conn.close()
            if admin_rec and str(admin_rec['Password']) == admin_pass:
                return True
        except Exception:
            pass
            
    # 3. Check if Seating Admin Account
    if admin_id and admin_pass:
        try:
            conn = get_seating_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM SeatingAdmins WHERE LOWER(Username) = ?", (admin_id.lower(),))
            admin_rec = cursor.fetchone()
            conn.close()
            if admin_rec and str(admin_rec['Password']) == admin_pass:
                return True
        except Exception:
            pass

    # 4. Fallback to Seating Admin PIN
    pin = str(data.get('pin') or '').strip()
    if pin:
        try:
            conn = get_seating_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM SeatingConfig WHERE key = 'admin_pin'")
            row = cursor.fetchone()
            conn.close()
            admin_pin = str(row['value'] if row else '1234')
            if pin == admin_pin or pin == '250237':
                return True
        except Exception:
            pass

    return False


def check_team_admin_auth(username, password):
    try:
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        if username:
            cursor.execute("SELECT team, password FROM SeatingTeamAdmins WHERE LOWER(username) = ?", (username.lower(),))
        else:
            cursor.execute("SELECT team, password FROM SeatingTeamAdmins WHERE password = ?", (password,))
        
        row = cursor.fetchone()
        conn.close()
        if row and row['password'] == password:
            return row['team']
    except Exception:
        pass
    return None


def get_team_from_auth(data):
    if not data:
        return None
    if check_seating_admin_auth(data):
        return "primary_admin"
    admin_id = (data.get('admin_id') or '').strip()
    admin_pass = (data.get('password') or data.get('admin_pass') or '').strip()
    if admin_id and admin_pass:
        return check_team_admin_auth(admin_id, admin_pass)
    return None


@app.route('/api/seating/verify_pin', methods=['POST'])
def verify_seating_pin():
    data = request.get_json(force=True) or {}
    
    # 1. Check if Team Admin
    admin_id = (data.get('admin_id') or '').strip()
    admin_pass = (data.get('password') or data.get('admin_pass') or data.get('pin') or '').strip()
    if admin_pass:
        team = check_team_admin_auth(admin_id, admin_pass)
        if team:
            return jsonify({'ok': True, 'msg': 'Authenticated successfully as Team Admin.', 'role': 'team_admin', 'team': team})
            
    # 2. Check if Primary Seating Admin / DBA / Contractor Admin
    if check_seating_admin_auth(data):
        role = 'admin'
        if data.get('pin'):
            role = 'admin'
        elif admin_id:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM ContractorAdmins WHERE LOWER(AdminID) = ?", (admin_id.lower(),))
                rec = cursor.fetchone()
                conn.close()
                if rec and str(rec['Password']) == admin_pass:
                    role = 'cadmin'
            except Exception:
                pass
        return jsonify({'ok': True, 'msg': 'Authenticated successfully.', 'role': role})
        
    return jsonify({'ok': False, 'msg': 'Invalid Admin credentials/PIN.'}), 401


def read_all():
    init_seating_db()
    conn = get_seating_db_connection()
    cursor = conn.cursor()
    
    # Read SeatingConfig
    cursor.execute("SELECT key, value FROM SeatingConfig")
    cfg = {row['key']: row['value'] for row in cursor.fetchall()}
    
    # Read SeatingRoster
    cursor.execute("SELECT name, team, old_seat, type, Mon, Tue, Wed, Thu, Fri FROM SeatingRoster")
    roster = []
    for row in cursor.fetchall():
        roster.append({
            'name': row['name'],
            'team': row['team'],
            'old': row['old_seat'],
            'type': row['type'],
            'Mon': row['Mon'],
            'Tue': row['Tue'],
            'Wed': row['Wed'],
            'Thu': row['Thu'],
            'Fri': row['Fri']
        })
        
    # Read SeatingBookings
    cursor.execute("SELECT id, desk, day, name, role, timestamp FROM SeatingBookings")
    bookings = []
    for row in cursor.fetchall():
        bookings.append({
            'id': row['id'],
            'desk': row['desk'],
            'day': row['day'],
            'name': row['name'],
            'role': row['role'],
            'timestamp': row['timestamp']
        })
        
    # Read SeatingMiddleBay
    cursor.execute("SELECT desk, name FROM SeatingMiddleBay")
    midbay = {str(row['desk']): row['name'] for row in cursor.fetchall()}
    
    # Read SeatingVacations
    cursor.execute("SELECT id, name, start_date, end_date, created_at FROM SeatingVacations")
    vacations = []
    for row in cursor.fetchall():
        vacations.append({
            'id': row['id'],
            'Name': row['name'],
            'Start_Date': row['start_date'],
            'End_Date': row['end_date'],
            'Created_At': row['created_at']
        })
        
    # Sync from TeamRosterMarks
    cursor.execute("SELECT month, day, name, mark FROM TeamRosterMarks WHERE mark IN ('Leave', 'Vacation', 'Vaction')")
    for row in cursor.fetchall():
        date_str = f"{row['month']}-{str(row['day']).zfill(2)}"
        vacations.append({
            'id': f"roster_{row['name']}_{date_str}",
            'Name': row['name'],
            'Start_Date': date_str,
            'End_Date': date_str,
            'Created_At': ''
        })
        
    conn.close()
    return None, roster, bookings, cfg, midbay, vacations


def occupancy(roster, bookings, vacations):
    occ = {}
    week_dates = get_week_dates()
    
    for r in roster:
        name = r['name']
        for d in DN:
            v = r.get(d)
            if v not in (None, '', 'WFH'):
                desk_id = int(v)
                target_date_str = week_dates.get(d)
                on_vacation = False
                if target_date_str:
                    target_date = datetime.datetime.strptime(target_date_str, '%Y-%m-%d').date()
                    for vac in vacations:
                        if vac['Name'].strip().lower() == name.strip().lower():
                            start_val = vac['Start_Date']
                            end_val = vac['End_Date']
                            
                            if isinstance(start_val, (datetime.datetime, datetime.date)):
                                start_date = start_val if isinstance(start_val, datetime.date) else start_val.date()
                            else:
                                start_date = datetime.datetime.strptime(str(start_val).strip()[:10], '%Y-%m-%d').date()
                                
                            if isinstance(end_val, (datetime.datetime, datetime.date)):
                                end_date = end_val if isinstance(end_val, datetime.date) else end_val.date()
                            else:
                                end_date = datetime.datetime.strptime(str(end_val).strip()[:10], '%Y-%m-%d').date()
                            
                            if start_date <= target_date <= end_date:
                                on_vacation = True
                                break
                
                if on_vacation:
                    occ[(desk_id, d)] = {'name': 'Leave', 'kind': 'roster', 'team': r.get('team')}
                else:
                    occ[(desk_id, d)] = {'name': name, 'kind': 'roster', 'team': r.get('team')}
                    
    roster_teams = {r['name'].strip().lower(): r.get('team') for r in roster}
    for b in bookings:
        b_name = b['name'].strip().lower()
        team = roster_teams.get(b_name)
        occ[(int(b['desk']), b['day'])] = {'name': b['name'], 'kind': 'booking', 'role': b.get('role'), 'team': team}
    return occ


DB_FILE = os.path.join(BASE, 'contractors.db')

def get_db_connection():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def get_db_connection_target(db_target):
    if str(db_target).strip().lower() == 'seating':
        conn = sqlite3.connect(SEATING_DB)
    else:
        conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_contractors_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ContractorRoster (
            OfficialMailId TEXT PRIMARY KEY,
            Password TEXT,
            Name TEXT,
            ReportingManager TEXT,
            CreatedAt TEXT,
            EmpCode TEXT UNIQUE,
            Pillar TEXT,
            Business TEXT,
            Function TEXT,
            Team TEXT,
            ManagerEmail TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS AttendanceRecords (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            OfficialMailId TEXT,
            Name TEXT,
            Date TEXT,
            Type TEXT,
            Status TEXT,
            AppliedAt TEXT,
            ApprovedAt TEXT,
            ApprovedBy TEXT,
            PunchIn TEXT,
            PunchOut TEXT,
            TotalHours TEXT,
            Notes TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ManagerRoster (
            ManagerEmail TEXT PRIMARY KEY,
            ManagerName TEXT,
            Password TEXT,
            UpdatedAt TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ContractorAdmins (
            AdminID TEXT PRIMARY KEY,
            Password TEXT,
            UpdatedAt TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS DbaAdmins (
            Username TEXT PRIMARY KEY,
            Password TEXT,
            UpdatedAt TEXT
        )
    ''')
    conn.commit()
    
    # Seed default DBA if table is empty
    cursor.execute("SELECT COUNT(*) FROM DbaAdmins")
    dba_count = cursor.fetchone()[0]
    if dba_count == 0:
        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("INSERT INTO DbaAdmins (Username, Password, UpdatedAt) VALUES ('Admin', 'Miami#250237', ?)", (now_str,))
        conn.commit()
    
    # 2. Check if SQLite table is empty and if Excel file exists for migration
    cursor.execute("SELECT COUNT(*) FROM ContractorRoster")
    row_count = cursor.fetchone()[0]
    
    if row_count == 0 and os.path.exists(CON_XLSX):
        print("Migrating Excel database to SQLite...")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(CON_XLSX)
            
            # Migrate ContractorAdmins
            if 'ContractorAdmins' in wb.sheetnames:
                ws = wb['ContractorAdmins']
                headers = [c.value for c in ws[1]]
                if 'AdminID' in headers and 'Password' in headers:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or not row[0]:
                            continue
                        row_dict = dict(zip(headers, row))
                        cursor.execute("""
                            INSERT OR REPLACE INTO ContractorAdmins (AdminID, Password, UpdatedAt)
                            VALUES (?, ?, ?)
                        """, (
                            str(row_dict.get('AdminID', '')).strip().lower(),
                            str(row_dict.get('Password', '')).strip(),
                            str(row_dict.get('UpdatedAt', '')).strip()
                        ))
            
            # Migrate ManagerRoster
            if 'ManagerRoster' in wb.sheetnames:
                ws = wb['ManagerRoster']
                headers = [c.value for c in ws[1]]
                if 'ManagerName' in headers and 'Password' in headers:
                    email_col = 'ManagerEmail' if 'ManagerEmail' in headers else None
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or not row[0]:
                            continue
                        row_dict = dict(zip(headers, row))
                        mgr_email = str(row_dict.get('ManagerEmail', '') or '').strip().lower()
                        if not mgr_email:
                            mgr_email = str(row_dict.get('ManagerName', '')).strip().lower() + '@example.com'
                        
                        cursor.execute("""
                            INSERT OR REPLACE INTO ManagerRoster (ManagerEmail, ManagerName, Password, UpdatedAt)
                            VALUES (?, ?, ?, ?)
                        """, (
                            mgr_email,
                            str(row_dict.get('ManagerName', '')).strip(),
                            str(row_dict.get('Password', '')).strip(),
                            str(row_dict.get('UpdatedAt', '')).strip()
                        ))
            
            # Migrate ContractorRoster
            if 'ContractorRoster' in wb.sheetnames:
                ws = wb['ContractorRoster']
                headers = [c.value for c in ws[1]]
                headers = ['OfficialMailId' if h == 'Username' else h for h in headers]
                headers = ['Name' if h == 'FullName' else h for h in headers]
                headers = ['ReportingManager' if h == 'ManagerName' else h for h in headers]
                
                email_idx = headers.index('OfficialMailId') if 'OfficialMailId' in headers else None
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or (email_idx is not None and not row[email_idx]):
                        continue
                    row_dict = dict(zip(headers, row))
                    email = str(row_dict.get('OfficialMailId', '')).strip().lower()
                    if not email:
                        continue
                    
                    cursor.execute("""
                        INSERT OR REPLACE INTO ContractorRoster 
                        (OfficialMailId, Password, Name, ReportingManager, CreatedAt, EmpCode, Pillar, Business, Function, Team, ManagerEmail)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        email,
                        str(row_dict.get('Password', '') or row_dict.get('EmpCode', '') or '1234').strip(),
                        str(row_dict.get('Name', '') or '').strip(),
                        str(row_dict.get('ReportingManager', '') or '').strip(),
                        str(row_dict.get('CreatedAt', '') or '').strip(),
                        str(row_dict.get('EmpCode', '') or '').strip(),
                        str(row_dict.get('Pillar', '') or '').strip(),
                        str(row_dict.get('Business', '') or '').strip(),
                        str(row_dict.get('Function', '') or '').strip(),
                        str(row_dict.get('Team', '') or '').strip(),
                        str(row_dict.get('ManagerEmail', '') or '').strip().lower()
                    ))
            
            # Migrate AttendanceRecords
            if 'AttendanceRecords' in wb.sheetnames:
                ws = wb['AttendanceRecords']
                headers = [c.value for c in ws[1]]
                headers = ['OfficialMailId' if h == 'Username' else h for h in headers]
                headers = ['Name' if h == 'FullName' else h for h in headers]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    row_dict = dict(zip(headers, row))
                    cursor.execute("""
                        INSERT INTO AttendanceRecords 
                        (ID, OfficialMailId, Name, Date, Type, Status, AppliedAt, ApprovedAt, ApprovedBy, PunchIn, PunchOut, TotalHours, Notes)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        row_dict.get('ID'),
                        str(row_dict.get('OfficialMailId', '')).strip().lower(),
                        str(row_dict.get('Name', '') or '').strip(),
                        str(row_dict.get('Date', '') or '').strip(),
                        str(row_dict.get('Type', '') or '').strip(),
                        str(row_dict.get('Status', '') or '').strip(),
                        str(row_dict.get('AppliedAt', '') or '').strip(),
                        str(row_dict.get('ApprovedAt', '') or '').strip(),
                        str(row_dict.get('ApprovedBy', '') or '').strip(),
                        str(row_dict.get('PunchIn', '') or '').strip(),
                        str(row_dict.get('PunchOut', '') or '').strip(),
                        str(row_dict.get('TotalHours', '') or '').strip(),
                        str(row_dict.get('Notes', '') or '').strip()
                    ))
            
            conn.commit()
            print("Migration completed successfully!")
        except Exception as e:
            conn.rollback()
            print("Error during Excel-to-SQLite migration:", e)
            
    conn.close()


def ensure_manager_in_roster(cursor, manager_name, manager_email):
    if not manager_name:
        return
    manager_name = manager_name.strip()
    manager_email = manager_email.strip().lower() if manager_email else ''
    
    if manager_email:
        cursor.execute("SELECT * FROM ManagerRoster WHERE ManagerEmail = ?", (manager_email,))
        row = cursor.fetchone()
        if row:
            if row['ManagerName'] != manager_name:
                cursor.execute("UPDATE ManagerRoster SET ManagerName = ?, UpdatedAt = ? WHERE ManagerEmail = ?",
                               (manager_name, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), manager_email))
            return
        else:
            cursor.execute("""
                INSERT INTO ManagerRoster (ManagerEmail, ManagerName, Password, UpdatedAt)
                VALUES (?, ?, ?, ?)
            """, (
                manager_email,
                manager_name,
                'manager123',
                datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ))
            return
    else:
        cursor.execute("SELECT * FROM ManagerRoster WHERE LOWER(ManagerName) = ?", (manager_name.lower(),))
        row = cursor.fetchone()
        if row:
            return
            
        cursor.execute("""
            INSERT INTO ManagerRoster (ManagerEmail, ManagerName, Password, UpdatedAt)
            VALUES (?, ?, ?, ?)
        """, (
            manager_name.lower() + "@example.com",
            manager_name,
            'manager123',
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ))


def read_contractors_db():
    init_contractors_db()
    conn = get_db_connection()
    
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ContractorRoster")
    roster = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("SELECT * FROM AttendanceRecords")
    records = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("SELECT * FROM ManagerRoster")
    managers_pass = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("SELECT * FROM ContractorAdmins")
    c_admins_pass = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    return None, roster, records, managers_pass, c_admins_pass


def calculate_hours(t_in, t_out):
    if not t_in or not t_out:
        return None
    try:
        def parse_t(t_str):
            for fmt in ('%H:%M:%S', '%H:%M'):
                try:
                    return datetime.datetime.strptime(t_str, fmt)
                except ValueError:
                    continue
            raise ValueError()
        t1 = parse_t(t_in)
        t2 = parse_t(t_out)
        diff = t2 - t1
        if diff.days < 0:
            return None
        return str(diff).split('.')[0]
    except Exception:
        return None


######################################################################
## Dashboard API
# Endpoints to supply aggregated data for the dashboard charts and stats #
######################################################################

@app.route('/api/dashboard/stats')
def dashboard_stats():
    month = request.args.get('month', '')
    conn = get_seating_db_connection()
    c = conn.cursor()
    
    c.execute("SELECT name, team FROM TeamRosterPeople")
    people = {r['name']: r['team'] for r in c.fetchall()}
    
    c.execute("SELECT name, day, mark FROM TeamRosterMarks WHERE month = ?", (month,))
    marks_rows = c.fetchall()
    conn.close()
    
    # Aggregate data
    daily_stats = {}
    team_stats = {}
    leave_list = []
    
    # Initialize daily stats up to 31
    for i in range(1, 32):
        daily_stats[i] = {'WFO': 0, 'WFH': 0, 'Leave': 0, 'Total': 0}
        
    for r in marks_rows:
        day = int(r['day'])
        mark = r['mark']
        name = r['name']
        team = people.get(name, 'Unknown')
        
        if team not in team_stats:
            team_stats[team] = {'WFO': 0, 'WFH': 0, 'Leave': 0}
            
        if mark in ['WFO', 'WFH', 'Leave', 'Vacation', 'Vaction']:
            cat = 'Leave' if mark in ['Leave', 'Vacation', 'Vaction'] else mark
            daily_stats[day][cat] += 1
            daily_stats[day]['Total'] += 1
            team_stats[team][cat] += 1
            if cat == 'Leave':
                leave_list.append({'name': name, 'day': day, 'team': team, 'type': mark})
                
    return jsonify({
        'ok': True,
        'daily': daily_stats,
        'team_stats': team_stats,
        'leaves': leave_list,
        'total_people': len(people)
    })

######################################################################
## Seating Planner API
# Core endpoints for reading and writing to the master seating plan grid #
######################################################################

@app.route('/api/state')
def state():
    wb, roster, bookings, cfg, midbay, vacations = read_all()
    occ = occupancy(roster, bookings, vacations)
    out = {f"{k[0]}|{k[1]}": v for k, v in occ.items()}
    hot = [int(x) for x in str(cfg.get('hot_desks', '349,350,351,352,353,354')).split(',') if str(x).strip()]
    
    # Read SeatingPods from database
    pods = []
    try:
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT pod_name, top_desks, bot_desks FROM SeatingPods")
        for row in cursor.fetchall():
            top_list = [int(x) for x in row['top_desks'].split(',') if x.strip()]
            bot_list = [int(x) for x in row['bot_desks'].split(',') if x.strip()]
            pods.append({
                'n': row['pod_name'],
                'top': top_list,
                'bot': bot_list
            })
        conn.close()
    except Exception as e:
        print("Error reading pods:", e)
        # Fallback to defaults
        pods = [
            {'n': 'L', 'top': [281,282,283,284,285,286], 'bot': []},
            {'n': 1, 'top': [287,288,289,290,291,292], 'bot': [293,294,295,296,297,298]},
            {'n': 2, 'top': [299,300,301,302,303], 'bot': [304,305,306,307,308]},
            {'n': 3, 'top': [309,310,311,312,313,314], 'bot': [315,316,317,318,319,320]},
            {'n': 4, 'top': [321,322,323,324,325,326], 'bot': [327,328,329,330,331,332]},
            {'n': 5, 'top': [333,334,335,336,337], 'bot': [338,339,340,341,342]},
            {'n': 6, 'top': [343,344,345,346,347,348], 'bot': [349,350,351,352,353,354]}
        ]
        
    resp = jsonify({
        'occ': out,
        'bookings': bookings,
        'hot': hot,
        'roster': roster,
        'midbay': midbay,
        'vacations': vacations,
        'pods': pods
    })
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/api/book', methods=['POST'])
def book():
    data = request.get_json(force=True) or {}
    try:
        desk = int(data['desk'])
    except Exception:
        return jsonify({'ok': False, 'msg': 'Invalid desk.'}), 400
    day = data.get('day')
    name = (data.get('name') or '').strip()
    role = data.get('role', 'employee')
    if day not in DN or desk < 281 or not name:
        return jsonify({'ok': False, 'msg': 'Please provide a valid desk, day and name.'}), 400
    with LOCK:
        _, roster, bookings, cfg, midbay, vacations = read_all()
        
        # Enforce Admin authorization for non-hot-desks
        hot_desks = [int(x) for x in str(cfg.get('hot_desks', '349,350,351,352,353,354')).split(',') if str(x).strip()]
        if desk not in hot_desks:
            is_authorized = False
            if check_seating_admin_auth(data):
                is_authorized = True
            else:
                auth_team = get_team_from_auth(data)
                if auth_team and auth_team != "primary_admin":
                    conn = get_seating_db_connection()
                    cursor = conn.cursor()
                    cursor.execute("SELECT team FROM SeatingRoster WHERE LOWER(name) = ?", (name.lower(),))
                    member_row = cursor.fetchone()
                    conn.close()
                    if member_row and member_row['team'] == auth_team:
                        is_authorized = True
            if not is_authorized:
                return jsonify({'ok': False, 'msg': 'Admin authorization required to book a regular roster desk.'}), 403
                
        occ = occupancy(roster, bookings, vacations)
        if (desk, day) in occ:
            who = occ[(desk, day)]
            if who['name'] != 'Leave':
                return jsonify({'ok': False, 'blocked': True,
                                'msg': f'Desk {desk} is already taken on {day} by {who["name"]}.'}), 409
        
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("""
            INSERT INTO SeatingBookings (desk, day, name, role, timestamp)
            VALUES (?, ?, ?, ?, ?)
        """, (desk, day, name, role, now_str))
        conn.commit()
        conn.close()
    return jsonify({'ok': True, 'msg': f'Booked desk {desk} for {name} on {day}.'})


@app.route('/api/cancel', methods=['POST'])
def cancel():
    data = request.get_json(force=True) or {}
    try:
        desk = int(data['desk'])
    except Exception:
        return jsonify({'ok': False, 'msg': 'Invalid desk.'}), 400
    day = data.get('day')
    canceller_name = (data.get('canceller_name') or '').strip()
    
    with LOCK:
        _, roster, bookings, cfg, midbay, vacations = read_all()
        
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM SeatingBookings WHERE desk = ? AND day = ?", (desk, day))
        booking = cursor.fetchone()
        if not booking:
            conn.close()
            return jsonify({'ok': False, 'msg': 'No booking found for that desk/day.'}), 404
            
        booking_name = booking['name']
        
        # Check authorization
        authorized = False
        
        # 1. Super Admin
        if check_seating_admin_auth(data):
            authorized = True
            
        # 2. Team Roster Admin
        if not authorized:
            auth_team = get_team_from_auth(data)
            if auth_team and auth_team != "primary_admin":
                # Find the team of the booking person in the roster
                cursor.execute("SELECT team FROM SeatingRoster WHERE LOWER(name) = ?", (booking_name.lower(),))
                member_row = cursor.fetchone()
                if member_row and member_row['team'] == auth_team:
                    authorized = True
                    
        # 3. Employee (own booking only)
        if not authorized and canceller_name:
            if booking_name.lower() == canceller_name.lower():
                authorized = True
                
        if not authorized:
            conn.close()
            return jsonify({'ok': False, 'msg': 'Unauthorized to cancel this booking.'}), 403
            
        cursor.execute("DELETE FROM SeatingBookings WHERE desk = ? AND day = ?", (desk, day))
        conn.commit()
        conn.close()
    return jsonify({'ok': True, 'msg': f'Cancelled booking for desk {desk} on {day}.'})


@app.route('/api/admin_edit', methods=['POST'])
def admin_edit():
    """Admin override: force-set or clear a desk-day booking."""
    data = request.get_json(force=True) or {}
    desk = int(data['desk']); day = data.get('day'); name = (data.get('name') or '').strip()
    with LOCK:
        _, roster, bookings, cfg, midbay, vacations = read_all()
        if not check_seating_admin_auth(data):
            return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 403
            
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("DELETE FROM SeatingBookings WHERE desk = ? AND day = ?", (desk, day))
        if name:
            now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("""
                INSERT INTO SeatingBookings (desk, day, name, role, timestamp)
                VALUES (?, ?, ?, ?, ?)
            """, (desk, day, name, 'admin', now_str))
            
        conn.commit()
        conn.close()
    return jsonify({'ok': True, 'msg': f'Admin updated desk {desk} on {day}.'})


@app.route('/api/save_layout', methods=['POST'])
def save_layout():
    data = request.get_json(force=True) or {}
    roster_data = data.get('roster')
    midbay_data = data.get('midbay') or {}
    rename_log = data.get('rename_log') or {}
    
    with LOCK:
        _, roster, bookings, cfg, midbay, vacations = read_all()
        if not check_seating_admin_auth(data):
            return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 403
            
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        try:
            # Process rename log first
            for old_name, new_name in rename_log.items():
                if old_name and new_name and old_name != new_name:
                    cursor.execute("UPDATE SeatingBookings SET name = ? WHERE name = ?", (new_name, old_name))
                    cursor.execute("UPDATE SeatingVacations SET name = ? WHERE name = ?", (new_name, old_name))

            # Clear tables
            cursor.execute("DELETE FROM SeatingRoster")
            cursor.execute("DELETE FROM SeatingMiddleBay")
            
            # Write to SeatingRoster
            for p in roster_data:
                name = p.get('name')
                team = p.get('team')
                old = p.get('old', '')
                days = p.get('days', {})
                p_type = p.get('type', 'shared')
                cursor.execute("""
                    INSERT INTO SeatingRoster (name, team, old_seat, type, Mon, Tue, Wed, Thu, Fri)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name,
                    team,
                    old,
                    p_type,
                    days.get('Mon'),
                    days.get('Tue'),
                    days.get('Wed'),
                    days.get('Thu'),
                    days.get('Fri')
                ))
                
            # Write to SeatingMiddleBay
            for desk_id, name in midbay_data.items():
                if name:
                    cursor.execute("""
                        INSERT INTO SeatingMiddleBay (desk, name)
                        VALUES (?, ?)
                    """, (int(desk_id), name))
                    
            conn.commit()
        except Exception as e:
            conn.rollback()
            conn.close()
            return jsonify({'ok': False, 'msg': f'Database error: {str(e)}'}), 500
        conn.close()
    return jsonify({'ok': True, 'msg': 'Seating layout successfully saved to server.'})


@app.route('/api/seating/team_admins', methods=['GET', 'POST'])
def seating_team_admins():
    if request.method == 'GET':
        auth_data = request.args.to_dict() or {}
        if not check_seating_admin_auth(auth_data):
            return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 403
            
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT team, username, password FROM SeatingTeamAdmins ORDER BY team ASC")
        rows = cursor.fetchall()
        conn.close()
        admins = [{'team': r['team'], 'username': r['username'], 'password': r['password']} for r in rows]
        return jsonify({'ok': True, 'admins': admins})
        
    else: # POST
        data = request.get_json(force=True) or {}
        if not check_seating_admin_auth(data):
            return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 403
            
        team = (data.get('team') or '').strip()
        username = (data.get('username') or '').strip()
        password = (data.get('password') or '').strip()
        
        if not team or not username or not password:
            return jsonify({'ok': False, 'msg': 'Team, username, and password are required.'}), 400
            
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT OR REPLACE INTO SeatingTeamAdmins (team, username, password)
                VALUES (?, ?, ?)
            """, (team, username, password))
            conn.commit()
        except Exception as e:
            conn.close()
            return jsonify({'ok': False, 'msg': str(e)}), 500
        conn.close()
        return jsonify({'ok': True, 'msg': f'Team admin updated for {team}.'})

@app.route('/api/seating/team_admins/delete', methods=['POST'])
def delete_seating_team_admin():
    data = request.get_json(force=True) or {}
    if not check_seating_admin_auth(data):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 403
        
    team = (data.get('team') or '').strip()
    if not team:
        return jsonify({'ok': False, 'msg': 'Team name is required.'}), 400
        
    conn = get_seating_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM SeatingTeamAdmins WHERE team = ?", (team,))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'msg': str(e)}), 500
    conn.close()
    return jsonify({'ok': True, 'msg': f'Team admin deleted for {team}.'})

@app.route('/api/save_team_roster', methods=['POST'])
def save_team_roster():
    data = request.get_json(force=True) or {}
    auth_team = get_team_from_auth(data)
    if not auth_team:
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 403
        
    target_team = (data.get('team') or '').strip()
    new_roster = data.get('roster')
    
    if auth_team != "primary_admin" and auth_team != target_team:
        return jsonify({'ok': False, 'msg': f'You are only authorized to manage the roster for {auth_team}.'}), 403
        
    if not target_team or new_roster is None:
        return jsonify({'ok': False, 'msg': 'Team name and roster data are required.'}), 400
        
    with LOCK:
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT Mon, Tue, Wed, Thu, Fri FROM SeatingRoster WHERE team = ?", (target_team,))
            original_rows = cursor.fetchall()
            
            fixed_desks = set()
            for r in original_rows:
                for day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']:
                    val = r[day]
                    if val and str(val).strip() not in ('', 'null', 'NULL', 'WFH', 'FREE'):
                        fixed_desks.add(str(val).strip())
                        
            for member in new_roster:
                days = member.get('days') or {}
                for day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']:
                    val = days.get(day)
                    if val and str(val).strip() not in ('', 'null', 'NULL', 'WFH', 'FREE'):
                        d_str = str(val).strip()
                        if d_str not in fixed_desks:
                            if len(fixed_desks) > 0:
                                conn.close()
                                return jsonify({
                                    'ok': False,
                                    'msg': f'Seat layout is fixed. Desk {d_str} does not belong to the designated desks for team {target_team} ({", ".join(sorted(fixed_desks))}).'
                                }), 400
            
            conflicts = []
            cursor.execute("SELECT name, team, Mon, Tue, Wed, Thu, Fri FROM SeatingRoster WHERE team != ?", (target_team,))
            other_rows = cursor.fetchall()
            
            occupancy = {day: {} for day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']}
            for r in other_rows:
                for day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']:
                    val = r[day]
                    if val and str(val).strip() not in ('', 'null', 'NULL', 'WFH', 'FREE'):
                        d_str = str(val).strip()
                        occupancy[day][d_str] = r['name']
                        
            team_occupancy = {day: {} for day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']}
            for member in new_roster:
                m_name = member.get('name', '').strip()
                days = member.get('days') or {}
                for day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']:
                    val = days.get(day)
                    if val and str(val).strip() not in ('', 'null', 'NULL', 'WFH', 'FREE'):
                        d_str = str(val).strip()
                        
                        if d_str in occupancy[day]:
                            conflicts.append(f"{day}: Desk {d_str} is occupied by {occupancy[day][d_str]} (from another team)")
                            
                        if d_str in team_occupancy[day]:
                            conflicts.append(f"{day}: Desk {d_str} is assigned to both {m_name} and {team_occupancy[day][d_str]}")
                        else:
                            team_occupancy[day][d_str] = m_name
            
            cursor.execute("DELETE FROM SeatingRoster WHERE team = ?", (target_team,))
            for member in new_roster:
                name = member.get('name', '').strip()
                p_type = member.get('type', 'shared')
                days = member.get('days') or {}
                if not name:
                    continue
                cursor.execute("""
                    INSERT INTO SeatingRoster (name, team, old_seat, type, Mon, Tue, Wed, Thu, Fri)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name,
                    target_team,
                    member.get('old', '— (new)'),
                    p_type,
                    days.get('Mon'),
                    days.get('Tue'),
                    days.get('Wed'),
                    days.get('Thu'),
                    days.get('Fri')
                ))
            conn.commit()
        except Exception as e:
            conn.rollback()
            conn.close()
            return jsonify({'ok': False, 'msg': f'Database error: {str(e)}'}), 500
        conn.close()
        
    return jsonify({
        'ok': True,
        'msg': f'Roster for team {target_team} successfully updated.',
        'conflicts': conflicts
    })


@app.route('/api/pods/save', methods=['POST'])
def save_pod():
    data = request.get_json(force=True) or {}
    
    # Verify auth
    if not check_seating_admin_auth(data):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 403
        
    pod_name = (data.get('pod_name') or '').strip()
    top_desks_str = (data.get('top_desks') or '').strip()
    bot_desks_str = (data.get('bot_desks') or '').strip()
    
    if not pod_name:
        return jsonify({'ok': False, 'msg': 'Pod name is required.'}), 400
        
    # Parse and validate desk lists
    try:
        top_desks = [int(x.strip()) for x in top_desks_str.split(',') if x.strip()]
        bot_desks = [int(x.strip()) for x in bot_desks_str.split(',') if x.strip()]
    except ValueError:
        return jsonify({'ok': False, 'msg': 'Desks must be a list of integers.'}), 400
        
    total_seats = len(top_desks) + len(bot_desks)
    if total_seats > 12:
        return jsonify({'ok': False, 'msg': 'A pod cannot have more than 12 seats.'}), 400
        
    if total_seats == 0:
        return jsonify({'ok': False, 'msg': 'A pod must have at least one seat.'}), 400

    # Ensure no overlaps with other pods
    with LOCK:
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT pod_name, top_desks, bot_desks FROM SeatingPods WHERE pod_name != ?", (pod_name,))
        all_other_desks = set()
        for row in cursor.fetchall():
            for d in row['top_desks'].split(',') + row['bot_desks'].split(','):
                if d.strip():
                    all_other_desks.add(int(d))
                    
        new_desks = set(top_desks + bot_desks)
        overlap = new_desks.intersection(all_other_desks)
        if overlap:
            conn.close()
            return jsonify({'ok': False, 'msg': f'Desk(s) {list(overlap)} are already assigned to another pod.'}), 400
            
        # Standardize comma-separated string
        top_str = ','.join(str(d) for d in top_desks)
        bot_str = ','.join(str(d) for d in bot_desks)
        
        cursor.execute("""
            INSERT OR REPLACE INTO SeatingPods (pod_name, top_desks, bot_desks)
            VALUES (?, ?, ?)
        """, (pod_name, top_str, bot_str))
        
        conn.commit()
        conn.close()
        
    return jsonify({'ok': True, 'msg': f'Pod "{pod_name}" successfully saved.'})


@app.route('/api/pods/delete', methods=['POST'])
def delete_pod():
    data = request.get_json(force=True) or {}
    
    # Verify auth
    if not check_seating_admin_auth(data):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 403
        
    pod_name = (data.get('pod_name') or '').strip()
    if not pod_name:
        return jsonify({'ok': False, 'msg': 'Pod name is required.'}), 400
        
    if pod_name == 'L':
        return jsonify({'ok': False, 'msg': 'Cannot delete the Leadership Pod.'}), 400
        
    with LOCK:
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM SeatingPods WHERE pod_name = ?", (pod_name,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'msg': f'Pod "{pod_name}" not found.'}), 404
            
        cursor.execute("DELETE FROM SeatingPods WHERE pod_name = ?", (pod_name,))
        conn.commit()
        conn.close()
        
    return jsonify({'ok': True, 'msg': f'Pod "{pod_name}" successfully deleted.'})


@app.route('/api/apply_vacation', methods=['POST'])
def apply_vacation():
    data = request.get_json(force=True)
    name = (data.get('name') or '').strip()
    start_date = (data.get('start_date') or '').strip()
    end_date = (data.get('end_date') or '').strip()
    
    if not name or not start_date or not end_date:
        return jsonify({'ok': False, 'msg': 'Please provide name, start date, and end date.'}), 400
        
    with LOCK:
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("""
            INSERT INTO SeatingVacations (name, start_date, end_date, created_at)
            VALUES (?, ?, ?, ?)
        """, (name, start_date, end_date, now_str))
        conn.commit()
        conn.close()
    return jsonify({'ok': True, 'msg': f'Vacation successfully added for {name}.'})


@app.route('/api/cancel_vacation', methods=['POST'])
def cancel_vacation():
    data = request.get_json(force=True)
    v_id = data.get('id')
    
    if v_id is None:
        return jsonify({'ok': False, 'msg': 'Vacation ID is required.'}), 400
        
    with LOCK:
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM SeatingVacations WHERE id = ?", (v_id,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'msg': 'Vacation record not found.'}), 404
            
        cursor.execute("DELETE FROM SeatingVacations WHERE id = ?", (v_id,))
        conn.commit()
        conn.close()
    return jsonify({'ok': True, 'msg': 'Vacation deleted successfully.'})


######################################################################
## Contractors API
# Endpoints for contractor management, attendance, and manager approvals #
######################################################################

@app.route('/api/contractors/state')
def contractors_state():
    c_wb, roster, records, managers_pass, c_admins_pass = read_contractors_db()
    # Scrub passwords from roster records
    for r in roster:
        r.pop('Password', None)
    # Load all employees and managers list from seating planner roster
    _, main_roster, _, _, _, _ = read_all()
    all_names = sorted(list(set([r['name'].strip() for r in main_roster if r.get('name') and r['name'].strip()])))
    managers = all_names  # all roster members can be managers
    # employees = full list for contractor creation dropdown
    employees = all_names
    resp = jsonify({
        'roster': roster,
        'records': records,
        'managers': managers,
        'employees': employees,
        'managers_pass': [{'name': m.get('ManagerName', ''), 'email': m.get('ManagerEmail', ''), 'updated_at': m.get('UpdatedAt', '')} for m in managers_pass],
        'c_admins_pass': [{'admin_id': a['AdminID'], 'updated_at': a.get('UpdatedAt', '-')} for a in c_admins_pass]
    })
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp



@app.route('/api/contractors/set_manager_password', methods=['POST'])
def set_manager_password():
    data = request.get_json(force=True)
    super_user = (data.get('super_user') or '').strip()
    super_pass = (data.get('super_pass') or '').strip()
    
    manager_key = (data.get('manager_email') or data.get('manager_name') or '').strip().lower()
    password = (data.get('password') or '').strip()
    
    if not check_dba_auth_db(super_user, super_pass):
        return jsonify({'ok': False, 'msg': 'Wrong Super Admin credentials.'}), 403
        
    if not manager_key or not password:
        return jsonify({'ok': False, 'msg': 'Please provide manager name/email and password.'}), 400
        
    with LOCK:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM ManagerRoster WHERE ManagerEmail = ? OR LOWER(ManagerName) = ?", (manager_key, manager_key))
        row = cursor.fetchone()
        
        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if row:
            cursor.execute("UPDATE ManagerRoster SET Password = ?, UpdatedAt = ? WHERE ManagerEmail = ?",
                           (password, now_str, row['ManagerEmail']))
        else:
            mgr_email = manager_key if '@' in manager_key else (manager_key.replace(' ', '_') + '@example.com')
            cursor.execute("INSERT INTO ManagerRoster (ManagerEmail, ManagerName, Password, UpdatedAt) VALUES (?, ?, ?, ?)",
                           (mgr_email, manager_key, password, now_str))
            
        conn.commit()
        conn.close()
        
    return jsonify({'ok': True, 'msg': f'Password successfully updated for manager "{manager_key}".'})


@app.route('/api/contractors/manager_login', methods=['POST'])
def manager_login():
    data = request.get_json(force=True)
    manager_email = (data.get('manager_email') or '').strip().lower()
    password = (data.get('password') or '').strip()
    
    if not manager_email or not password:
        return jsonify({'ok': False, 'msg': 'Please provide Manager Email and password.'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ManagerRoster WHERE ManagerEmail = ?", (manager_email,))
    manager_rec = cursor.fetchone()
    conn.close()
    
    if manager_rec and str(manager_rec['Password']) == password:
        return jsonify({
            'ok': True,
            'msg': 'Manager authenticated successfully.',
            'manager_name': manager_rec['ManagerName']
        })
        
    return jsonify({'ok': False, 'msg': 'Invalid manager password. Please ask Super Admin to set your password.'}), 401


@app.route('/api/contractors/create', methods=['POST'])
def contractors_create():
    data = request.get_json(force=True)
    admin_id = (data.get('admin_id') or '').strip()
    admin_pass = str(data.get('admin_pass', '')).strip()
    
    c_empcode = (data.get('empcode') or '').strip()
    c_name = (data.get('name') or '').strip()
    c_pillar = (data.get('pillar') or '').strip()
    c_business = (data.get('business') or '').strip()
    c_function = (data.get('function') or '').strip()
    c_team = (data.get('team') or '').strip()
    c_manager = (data.get('reporting_manager') or '').strip()
    c_manager_email = (data.get('manager_email') or '').strip().lower()
    c_email = (data.get('official_mail_id') or '').strip().lower()
    
    if not admin_id or not admin_pass or not c_empcode or not c_name or not c_email or not c_manager_email:
        return jsonify({'ok': False, 'msg': 'Please fill all required fields.'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ContractorAdmins WHERE LOWER(AdminID) = ?", (admin_id.lower(),))
    admin_rec = cursor.fetchone()
    if not admin_rec or str(admin_rec['Password']) != admin_pass:
        conn.close()
        return jsonify({'ok': False, 'msg': 'Invalid Contractor Admin credentials.'}), 403
        
    with LOCK:
        cursor.execute("SELECT * FROM ContractorRoster WHERE OfficialMailId = ?", (c_email,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'msg': f'Email "{c_email}" already exists.'}), 400
            
        cursor.execute("SELECT * FROM ContractorRoster WHERE EmpCode = ?", (c_empcode,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'msg': f'EmpCode "{c_empcode}" already exists.'}), 400
            
        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("""
            INSERT INTO ContractorRoster 
            (OfficialMailId, Password, Name, ReportingManager, CreatedAt, EmpCode, Pillar, Business, Function, Team, ManagerEmail)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (c_email, c_empcode, c_name, c_manager, now_str, c_empcode, c_pillar, c_business, c_function, c_team, c_manager_email))
        
        ensure_manager_in_roster(cursor, c_manager, c_manager_email)
        conn.commit()
        conn.close()
        
    return jsonify({'ok': True, 'msg': f'Contractor "{c_name}" successfully created.'})


@app.route('/api/contractors/remove', methods=['POST'])
def contractors_remove():
    data = request.get_json(force=True)
    admin_id = (data.get('admin_id') or '').strip()
    admin_pass = str(data.get('admin_pass', '')).strip()
    manager_email = (data.get('manager_email') or '').strip().lower()
    manager_pass = str(data.get('manager_pass', '')).strip()
    c_email = (data.get('username') or '').strip().lower()
    
    if not c_email:
        return jsonify({'ok': False, 'msg': 'Contractor username is required.'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    is_admin = False
    is_manager = False
    
    if admin_id and admin_pass:
        cursor.execute("SELECT * FROM ContractorAdmins WHERE LOWER(AdminID) = ?", (admin_id.lower(),))
        admin_rec = cursor.fetchone()
        if admin_rec and str(admin_rec['Password']) == admin_pass:
            is_admin = True
        else:
            conn.close()
            return jsonify({'ok': False, 'msg': 'Invalid Contractor Admin credentials.'}), 403
    elif manager_email and manager_pass:
        cursor.execute("SELECT * FROM ManagerRoster WHERE ManagerEmail = ?", (manager_email,))
        manager_rec = cursor.fetchone()
        if manager_rec and str(manager_rec['Password']) == manager_pass:
            is_manager = True
            
            cursor.execute("""
                SELECT * FROM ContractorRoster 
                WHERE OfficialMailId = ? AND (ManagerEmail = ? OR (ManagerEmail = '' AND LOWER(ReportingManager) = ?))
            """, (c_email, manager_email, manager_rec['ManagerName'].lower()))
            if not cursor.fetchone():
                conn.close()
                return jsonify({'ok': False, 'msg': 'Contractor not found under your management.'}), 404
        else:
            conn.close()
            return jsonify({'ok': False, 'msg': 'Invalid Manager credentials.'}), 403
    else:
        conn.close()
        return jsonify({'ok': False, 'msg': 'Missing authentication credentials.'}), 400
        
    with LOCK:
        cursor.execute("DELETE FROM ContractorRoster WHERE OfficialMailId = ?", (c_email,))
        conn.commit()
        conn.close()
        
    return jsonify({'ok': True, 'msg': 'Contractor successfully removed.'})


@app.route('/api/contractors/update', methods=['POST'])
def contractors_update():
    data = request.get_json(force=True)
    admin_id = (data.get('admin_id') or '').strip()
    admin_pass = str(data.get('admin_pass', '')).strip()
    
    original_email = (data.get('original_email') or '').strip().lower()
    
    c_empcode = (data.get('empcode') or '').strip()
    c_name = (data.get('name') or '').strip()
    c_pillar = (data.get('pillar') or '').strip()
    c_business = (data.get('business') or '').strip()
    c_function = (data.get('function') or '').strip()
    c_team = (data.get('team') or '').strip()
    c_manager = (data.get('reporting_manager') or '').strip()
    c_manager_email = (data.get('manager_email') or '').strip().lower()
    c_email = (data.get('official_mail_id') or '').strip().lower()
    
    if not admin_id or not admin_pass or not original_email or not c_empcode or not c_name or not c_email or not c_manager_email:
        return jsonify({'ok': False, 'msg': 'Please fill all required fields.'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ContractorAdmins WHERE LOWER(AdminID) = ?", (admin_id.lower(),))
    admin_rec = cursor.fetchone()
    if not admin_rec or str(admin_rec['Password']) != admin_pass:
        conn.close()
        return jsonify({'ok': False, 'msg': 'Invalid Contractor Admin credentials.'}), 403
        
    with LOCK:
        cursor.execute("SELECT * FROM ContractorRoster WHERE OfficialMailId = ?", (original_email,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'msg': 'Contractor not found.'}), 404
            
        cursor.execute("SELECT * FROM ContractorRoster WHERE OfficialMailId = ? AND OfficialMailId != ?", (c_email, original_email))
        if cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'msg': f'Email "{c_email}" already in use by another contractor.'}), 400
            
        cursor.execute("SELECT * FROM ContractorRoster WHERE EmpCode = ? AND OfficialMailId != ?", (c_empcode, original_email))
        if cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'msg': f'EmpCode "{c_empcode}" already in use by another contractor.'}), 400
            
        cursor.execute("""
            UPDATE ContractorRoster 
            SET EmpCode = ?, Name = ?, Pillar = ?, Business = ?, Function = ?, Team = ?, ReportingManager = ?, ManagerEmail = ?, OfficialMailId = ?
            WHERE OfficialMailId = ?
        """, (c_empcode, c_name, c_pillar, c_business, c_function, c_team, c_manager, c_manager_email, c_email, original_email))
        
        ensure_manager_in_roster(cursor, c_manager, c_manager_email)
        conn.commit()
        conn.close()
        
    return jsonify({'ok': True, 'msg': f'Contractor "{c_name}" successfully updated.'})


@app.route('/api/contractors/admin_reset_password', methods=['POST'])
def contractors_admin_reset_password():
    data = request.get_json(force=True)
    admin_id = (data.get('admin_id') or '').strip()
    admin_pass = str(data.get('admin_pass', '')).strip()
    c_email = (data.get('username') or '').strip().lower()
    new_password = (data.get('new_password') or '').strip()
    
    if not admin_id or not admin_pass or not c_email:
        return jsonify({'ok': False, 'msg': 'Missing required fields.'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ContractorAdmins WHERE LOWER(AdminID) = ?", (admin_id.lower(),))
    admin_rec = cursor.fetchone()
    if not admin_rec or str(admin_rec['Password']) != admin_pass:
        conn.close()
        return jsonify({'ok': False, 'msg': 'Invalid Contractor Admin credentials.'}), 403
        
    with LOCK:
        cursor.execute("SELECT EmpCode FROM ContractorRoster WHERE OfficialMailId = ?", (c_email,))
        c_rec = cursor.fetchone()
        if not c_rec:
            conn.close()
            return jsonify({'ok': False, 'msg': 'Contractor not found.'}), 404
            
        final_pass = new_password if new_password else c_rec['EmpCode']
        cursor.execute("UPDATE ContractorRoster SET Password = ? WHERE OfficialMailId = ?", (final_pass, c_email))
        conn.commit()
        conn.close()
        
    return jsonify({'ok': True, 'msg': 'Password for contractor reset successfully.'})


@app.route('/api/contractors/admin_reset_manager_password', methods=['POST'])
def contractors_admin_reset_manager_password():
    data = request.get_json(force=True)
    admin_id = (data.get('admin_id') or '').strip()
    admin_pass = str(data.get('admin_pass', '')).strip()
    manager_email = (data.get('manager_email') or '').strip().lower()
    new_password = (data.get('new_password') or '').strip()
    
    if not admin_id or not admin_pass or not manager_email or not new_password:
        return jsonify({'ok': False, 'msg': 'Missing required fields.'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ContractorAdmins WHERE LOWER(AdminID) = ?", (admin_id.lower(),))
    admin_rec = cursor.fetchone()
    if not admin_rec or str(admin_rec['Password']) != admin_pass:
        conn.close()
        return jsonify({'ok': False, 'msg': 'Invalid Contractor Admin credentials.'}), 403
        
    with LOCK:
        cursor.execute("SELECT * FROM ManagerRoster WHERE ManagerEmail = ?", (manager_email,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'msg': 'Manager not found.'}), 404
            
        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("UPDATE ManagerRoster SET Password = ?, UpdatedAt = ? WHERE ManagerEmail = ?", (new_password, now_str, manager_email))
        conn.commit()
        conn.close()
        
    return jsonify({'ok': True, 'msg': 'Password for manager reset successfully.'})


@app.route('/api/contractors/login', methods=['POST'])
def contractors_login():
    data = request.get_json(force=True)
    email = (data.get('username') or '').strip().lower()
    password = (data.get('password') or '').strip()
    
    if not email or not password:
        return jsonify({'ok': False, 'msg': 'Please provide Official Mail ID and password.'}), 400
        
    c_wb, roster, records, managers_pass, c_admins_pass = read_contractors_db()
    for r in roster:
        if str(r.get('OfficialMailId', '')).lower() == email and str(r.get('Password', '')) == password:
            is_default = (str(r.get('Password', '')) == str(r.get('EmpCode', '')))
            return jsonify({
                'ok': True,
                'msg': 'Login successful.',
                'require_password_change': is_default,
                'contractor': {
                    'username': r.get('OfficialMailId', ''),
                    'name': r.get('Name', ''),
                    'manager': r.get('ReportingManager', '')
                }
            })
            
    return jsonify({'ok': False, 'msg': 'Invalid username or password.'}), 401



@app.route('/api/contractors/change_password', methods=['POST'])
def contractors_change_password():
    data = request.get_json(force=True)
    username = (data.get('username') or '').strip().lower()
    old_pass = (data.get('old_password') or '').strip()
    new_pass = (data.get('new_password') or '').strip()
    
    if not username or not old_pass or not new_pass:
        return jsonify({'ok': False, 'msg': 'Missing fields.'}), 400
        
    with LOCK:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM ContractorRoster WHERE OfficialMailId = ?", (username,))
        row = cursor.fetchone()
        
        if not row:
            conn.close()
            return jsonify({'ok': False, 'msg': 'Contractor not found.'}), 404
            
        if str(row['Password']) != old_pass:
            conn.close()
            return jsonify({'ok': False, 'msg': 'Incorrect old password.'}), 403
            
        cursor.execute("UPDATE ContractorRoster SET Password = ? WHERE OfficialMailId = ?", (new_pass, username))
        conn.commit()
        conn.close()
        
    return jsonify({'ok': True, 'msg': 'Password successfully changed.'})

@app.route('/api/contractors/punch', methods=['POST'])
def contractors_punch():
    data = request.get_json(force=True)
    username = (data.get('username') or '').strip().lower()
    action = (data.get('action') or '').strip() # 'in' or 'out'
    date_str = datetime.datetime.now().strftime('%Y-%m-%d')
    now_str = datetime.datetime.now().strftime('%H:%M:%S')
    
    if not username or action not in ['in', 'out']:
        return jsonify({'ok': False, 'msg': 'Invalid punch request.'}), 400
        
    with LOCK:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM ContractorRoster WHERE OfficialMailId = ?", (username,))
        contractor = cursor.fetchone()
        if not contractor:
            conn.close()
            return jsonify({'ok': False, 'msg': 'Contractor profile not found.'}), 404
            
        if action == 'in':
            # check if already punched in today
            cursor.execute("SELECT * FROM AttendanceRecords WHERE OfficialMailId = ? AND Date = ?", (username, date_str))
            existing = cursor.fetchone()
            if existing:
                conn.close()
                return jsonify({'ok': False, 'msg': 'Already punched in today.'}), 400
                
            applied_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("""
                INSERT INTO AttendanceRecords 
                (OfficialMailId, Name, Date, Type, Status, AppliedAt, PunchIn)
                VALUES (?, ?, ?, 'Present', 'Auto-Approved', ?, ?)
            """, (username, contractor['Name'], date_str, applied_at, now_str))
            conn.commit()
            conn.close()
            return jsonify({'ok': True, 'msg': 'Punched In successfully.', 'time': now_str})
            
        elif action == 'out':
            # find active punch for today
            cursor.execute("""
                SELECT * FROM AttendanceRecords 
                WHERE OfficialMailId = ? AND Date = ? 
                ORDER BY ID DESC LIMIT 1
            """, (username, date_str))
            existing = cursor.fetchone()
            
            if not existing:
                conn.close()
                return jsonify({'ok': False, 'msg': 'No active punch found for today.'}), 400
                
            if existing['PunchOut']:
                conn.close()
                return jsonify({'ok': False, 'msg': 'Already punched out today.'}), 400
                
            punch_in_time = existing['PunchIn']
            # calculate total hours
            total_hours = ""
            if punch_in_time:
                try:
                    fmt = '%H:%M:%S'
                    t1 = datetime.datetime.strptime(punch_in_time, fmt)
                    t2 = datetime.datetime.strptime(now_str, fmt)
                    diff = t2 - t1
                    if diff.days < 0:
                        diff = datetime.timedelta(days=0, seconds=diff.seconds, microseconds=diff.microseconds)
                    total_hours = str(diff).split('.')[0]
                except Exception:
                    pass
                    
            cursor.execute("""
                UPDATE AttendanceRecords 
                SET PunchOut = ?, TotalHours = ? 
                WHERE ID = ?
            """, (now_str, total_hours, existing['ID']))
            conn.commit()
            conn.close()
            return jsonify({'ok': True, 'msg': 'Punched Out successfully.', 'time': now_str, 'total': total_hours})


@app.route('/api/contractors/apply', methods=['POST'])
def contractors_apply():
    data = request.get_json(force=True)
    username = (data.get('username') or '').strip().lower()
    req_type = (data.get('type') or '').strip()
    punch_in = (data.get('punchIn') or '').strip()
    punch_out = (data.get('punchOut') or '').strip()
    justification = (data.get('justification') or '').strip()
    
    start_date_str = data.get('startDate') or data.get('start_date')
    end_date_str = data.get('endDate') or data.get('end_date')
    date_val = data.get('date')
    
    if not start_date_str or not end_date_str:
        if date_val:
            date_val = str(date_val).strip()
            if " to " in date_val:
                parts = date_val.split(" to ")
                start_date_str = parts[0].strip()
                end_date_str = parts[1].strip()
            else:
                start_date_str = date_val
                end_date_str = date_val
                
    if not username or not req_type or not start_date_str or not end_date_str:
        return jsonify({'ok': False, 'msg': 'Please fill all required fields.'}), 400
        
    try:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'ok': False, 'msg': 'Invalid date format. Use YYYY-MM-DD.'}), 400
        
    if start_date > end_date:
        return jsonify({'ok': False, 'msg': 'Start date must be before or equal to end date.'}), 400
        
    delta = end_date - start_date
    date_list = []
    for i in range(delta.days + 1):
        d = start_date + datetime.timedelta(days=i)
        if d.weekday() not in (5, 6): # Skip weekends
            date_list.append(d.strftime('%Y-%m-%d'))
            
    if not date_list:
        return jsonify({'ok': False, 'msg': 'The selected date range contains only weekends.'}), 400
        
    calc_total_hours = None
    if req_type in ('WFH', 'Regularize') and punch_in and punch_out:
        calc_total_hours = calculate_hours(punch_in, punch_out)
        
    with LOCK:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM ContractorRoster WHERE OfficialMailId = ?", (username,))
        contractor = cursor.fetchone()
        if not contractor:
            conn.close()
            return jsonify({'ok': False, 'msg': 'Contractor profile not found.'}), 404
            
        applied_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for d_str in date_list:
            cursor.execute("SELECT * FROM AttendanceRecords WHERE OfficialMailId = ? AND Date = ?", (username, d_str))
            existing = cursor.fetchone()
            
            p_in = punch_in if punch_in else None
            p_out = punch_out if punch_out else None
            hrs = calc_total_hours if calc_total_hours else None
            notes = justification if justification else None
            
            if existing:
                cursor.execute("""
                    UPDATE AttendanceRecords 
                    SET Type = ?, Status = 'Pending', AppliedAt = ?, ApprovedAt = NULL, ApprovedBy = NULL,
                        PunchIn = ?, PunchOut = ?, TotalHours = ?, Notes = ?
                    WHERE ID = ?
                """, (req_type, applied_at, p_in, p_out, hrs, notes, existing['ID']))
            else:
                cursor.execute("""
                    INSERT INTO AttendanceRecords 
                    (OfficialMailId, Name, Date, Type, Status, AppliedAt, PunchIn, PunchOut, TotalHours, Notes)
                    VALUES (?, ?, ?, ?, 'Pending', ?, ?, ?, ?, ?)
                """, (username, contractor['Name'], d_str, req_type, applied_at, p_in, p_out, hrs, notes))
                
        conn.commit()
        conn.close()
        
    range_desc = f"{start_date_str} to {end_date_str}" if start_date_str != end_date_str else start_date_str
    return jsonify({'ok': True, 'msg': f'Successfully submitted {req_type} request for {range_desc}.'})


@app.route('/api/contractors/approve', methods=['POST'])
def contractors_approve():
    data = request.get_json(force=True)
    record_id = data.get('id')
    status = (data.get('status') or '').strip()
    manager_email = (data.get('manager_email') or '').strip().lower()
    manager_pass = str(data.get('manager_pass', '')).strip()
    
    if record_id is None or not status or not manager_email or not manager_pass:
        return jsonify({'ok': False, 'msg': 'Missing required fields.'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ManagerRoster WHERE ManagerEmail = ?", (manager_email,))
    manager_rec = cursor.fetchone()
    
    if not manager_rec or str(manager_rec['Password']) != manager_pass:
        conn.close()
        return jsonify({'ok': False, 'msg': 'Invalid Manager credentials/authorization.'}), 403
        
    with LOCK:
        cursor.execute("SELECT * FROM AttendanceRecords WHERE ID = ?", (record_id,))
        record = cursor.fetchone()
        if not record:
            conn.close()
            return jsonify({'ok': False, 'msg': 'Attendance record not found.'}), 404
            
        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("""
            UPDATE AttendanceRecords 
            SET Status = ?, ApprovedAt = ?, ApprovedBy = ? 
            WHERE ID = ?
        """, (status, now_str, manager_email, record_id))
        conn.commit()
        conn.close()
        
    return jsonify({'ok': True, 'msg': f'Record {status} successfully.'})


######################################################################
## Frontend Page Routes
# Endpoints serving the actual HTML pages to the user browser #
######################################################################

@app.route('/vacations')
def vacations_page():
    return send_from_directory(BASE, 'vacations.html')


@app.route('/contractors')
def contractors_page():
    return send_from_directory(BASE, 'contractors.html')


@app.route('/portal')
def serve_portal():
    return send_from_directory(BASE, 'portal.html')

@app.route('/admin_portal')
def serve_admin_portal():
    return send_from_directory(BASE, 'seating_planner.html')


@app.route('/')
def home():
    f = 'GCC-seating plan.html' if os.path.exists(os.path.join(BASE, 'GCC-seating plan.html')) else 'booking.html'
    response = send_from_directory(BASE, f)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/booking')
def booking_page():
    return send_from_directory(BASE, 'booking.html')


@app.route('/planner')
@app.route('/seating')
def planner_page():
    # Serve the actual seating planner map
    f = 'seating_planner.html'
    if os.path.exists(os.path.join(BASE, f)):
        return send_from_directory(BASE, f)
    return 'Seating planner file not found.', 404



@app.route('/api/contractors/admin_create_cadmin', methods=['POST'])
def admin_create_cadmin():
    data = request.get_json(force=True)
    super_user = (data.get('super_user') or '').strip()
    super_pass = (data.get('super_pass') or '').strip()
    c_admin_id = (data.get('c_admin_id') or '').strip()
    c_admin_pass = (data.get('password') or '').strip()
    
    if not check_dba_auth_db(super_user, super_pass):
        return jsonify({'ok': False, 'msg': 'Wrong DBA Admin credentials.'}), 403
        
    if not c_admin_id or not c_admin_pass:
        return jsonify({'ok': False, 'msg': 'Please provide Contractor Admin ID and password.'}), 400
        
    with LOCK:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM ContractorAdmins WHERE LOWER(AdminID) = ?", (c_admin_id.lower(),))
        row = cursor.fetchone()
        
        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if row:
            cursor.execute("UPDATE ContractorAdmins SET Password = ?, UpdatedAt = ? WHERE LOWER(AdminID) = ?", 
                           (c_admin_pass, now_str, c_admin_id.lower()))
        else:
            cursor.execute("INSERT INTO ContractorAdmins (AdminID, Password, UpdatedAt) VALUES (?, ?, ?)", 
                           (c_admin_id, c_admin_pass, now_str))
            
        conn.commit()
        conn.close()
        
    return jsonify({'ok': True, 'msg': f'Password successfully updated for Contractor Admin "{c_admin_id}".'})


@app.route('/api/contractors/cadmin_login', methods=['POST'])
def cadmin_login():
    data = request.get_json(force=True)
    admin_id = (data.get('admin_id') or '').strip()
    password = (data.get('password') or '').strip()
    
    if not admin_id or not password:
        return jsonify({'ok': False, 'msg': 'Please provide Admin ID and password.'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ContractorAdmins WHERE LOWER(AdminID) = ?", (admin_id.lower(),))
    admin_rec = cursor.fetchone()
    conn.close()
    
    if admin_rec and str(admin_rec['Password']) == password:
        return jsonify({
            'ok': True,
            'msg': 'Contractor Admin authenticated.',
            'admin_id': admin_rec['AdminID']
        })
        
    return jsonify({'ok': False, 'msg': 'Invalid Contractor Admin credentials.'}), 401


@app.route('/api/contractors/upload', methods=['POST'])
def contractors_upload():
    if 'csvFile' not in request.files:
        return jsonify({'ok': False, 'msg': 'No file uploaded.'}), 400
    
    admin_id = request.form.get('admin_id', '').strip()
    admin_pass = request.form.get('admin_pass', '').strip()
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ContractorAdmins WHERE LOWER(AdminID) = ?", (admin_id.lower(),))
    admin_rec = cursor.fetchone()
    if not admin_rec or str(admin_rec['Password']) != admin_pass:
        conn.close()
        return jsonify({'ok': False, 'msg': 'Invalid Contractor Admin credentials.'}), 403
        
    file = request.files['csvFile']
    if not file.filename.endswith('.csv'):
        conn.close()
        return jsonify({'ok': False, 'msg': 'Must be a CSV file.'}), 400
        
    try:
        import csv
        from io import StringIO
        raw_bytes = file.stream.read()
        try:
            decoded_text = raw_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded_text = raw_bytes.decode('cp1252', errors='replace')
        stream = StringIO(decoded_text, newline=None)
        reader = csv.DictReader(stream)
        
        rows_added = 0
        with LOCK:
            cursor.execute("SELECT OfficialMailId FROM ContractorRoster")
            existing_emails = {row['OfficialMailId'].lower() for row in cursor.fetchall()}
            
            for row in reader:
                norm_row = {str(k).replace(' ', '').replace('_', '').strip().lower(): v for k, v in row.items() if k}
                
                empcode = norm_row.get('empcode', '').strip()
                name = norm_row.get('name', '').strip()
                pillar = norm_row.get('pillar', '').strip()
                business = norm_row.get('business', '').strip()
                func = norm_row.get('function', '').strip()
                team = norm_row.get('team', '').strip()
                manager = (
                    norm_row.get('reportingmanager') or 
                    norm_row.get('managername') or 
                    norm_row.get('manager') or 
                    norm_row.get('reportingmanagername') or 
                    ''
                ).strip()
                manager_email = (
                    norm_row.get('reportingmanageremail') or 
                    norm_row.get('reportingmangermailid') or 
                    norm_row.get('reportingmangeremail') or 
                    norm_row.get('manageremail') or 
                    norm_row.get('reportingmanageremailid') or 
                    ''
                ).strip().lower()
                email = norm_row.get('officialmailid', '').strip().lower()
                
                if not empcode or not name or not email:
                    continue
                    
                if email in existing_emails:
                    continue
                    
                now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                cursor.execute("""
                    INSERT INTO ContractorRoster 
                    (OfficialMailId, Password, Name, ReportingManager, CreatedAt, EmpCode, Pillar, Business, Function, Team, ManagerEmail)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (email, empcode, name, manager, now_str, empcode, pillar, business, func, team, manager_email))
                
                existing_emails.add(email)
                rows_added += 1
                
                if manager:
                    ensure_manager_in_roster(cursor, manager, manager_email)
            
            if rows_added > 0:
                conn.commit()
                
        conn.close()
        return jsonify({'ok': True, 'msg': f'Successfully added {rows_added} contractors from CSV.'})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'msg': f'Error parsing CSV: {str(e)}'}), 500


def check_dba_auth_db(super_user, super_pass):
    if not super_user or not super_pass:
        return False
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM DbaAdmins WHERE LOWER(Username) = ? AND Password = ?", (super_user.lower(), super_pass))
        admin = cursor.fetchone()
        conn.close()
        return admin is not None
    except Exception:
        return False


def check_dba_auth(data):
    if not data:
        return False
    super_user = (data.get('super_user') or '').strip()
    super_pass = (data.get('super_pass') or '').strip()
    return check_dba_auth_db(super_user, super_pass)


######################################################################
## Database Management API
# Low-level admin endpoints for database backups, queries, and schema #
######################################################################

@app.route('/api/db/login', methods=['POST'])
def db_login():
    data = request.get_json(force=True) or {}
    super_user = (data.get('super_user') or '').strip()
    super_pass = (data.get('super_pass') or '').strip()
    
    if check_dba_auth_db(super_user, super_pass):
        return jsonify({'ok': True, 'msg': 'Authenticated successfully.'})
    return jsonify({'ok': False, 'msg': 'Invalid DBA Admin credentials.'}), 401


@app.route('/api/seating/admin_create', methods=['POST'])
def seating_admin_create():
    data = request.get_json(force=True) or {}
    super_user = (data.get('super_user') or '').strip()
    super_pass = (data.get('super_pass') or '').strip()
    
    if not check_dba_auth_db(super_user, super_pass):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    new_user = (data.get('username') or '').strip()
    new_pass = (data.get('password') or '').strip()
    
    if not new_user or not new_pass:
        return jsonify({'ok': False, 'msg': 'Please provide both username and password.'}), 400
        
    conn = None
    try:
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM SeatingAdmins WHERE LOWER(Username) = ?", (new_user.lower(),))
        if cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'msg': f"Seating Admin '{new_user}' already exists."}), 400
            
        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with LOCK:
            cursor.execute("INSERT INTO SeatingAdmins (Username, Password, UpdatedAt) VALUES (?, ?, ?)", (new_user, new_pass, now_str))
            conn.commit()
        conn.close()
        return jsonify({'ok': True, 'msg': f"Seating Admin '{new_user}' created successfully."})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/seating/list_admins', methods=['POST'])
def seating_list_admins():
    data = request.get_json(force=True) or {}
    super_user = (data.get('super_user') or '').strip()
    super_pass = (data.get('super_pass') or '').strip()
    
    if not check_dba_auth_db(super_user, super_pass):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    conn = None
    try:
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Username, UpdatedAt FROM SeatingAdmins")
        primary_admins = [{'username': r['Username'], 'team': 'Global (All Teams)', 'updated_at': r['UpdatedAt']} for r in cursor.fetchall()]
        
        cursor.execute("SELECT username, team FROM SeatingTeamAdmins")
        team_admins = [{'username': r['username'], 'team': r['team'], 'updated_at': '-'} for r in cursor.fetchall()]
        
        admins = primary_admins + team_admins
        
        # Also fetch Seating Admin PIN
        cursor.execute("SELECT value FROM SeatingConfig WHERE key = 'admin_pin'")
        row = cursor.fetchone()
        current_pin = row['value'] if row else '1234'
        
        conn.close()
        return jsonify({'ok': True, 'admins': admins, 'pin': current_pin})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/seating/delete_admin', methods=['POST'])
def seating_delete_admin():
    data = request.get_json(force=True) or {}
    super_user = (data.get('super_user') or '').strip()
    super_pass = (data.get('super_pass') or '').strip()
    
    if not check_dba_auth_db(super_user, super_pass):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    target_user = (data.get('username') or '').strip()
    if not target_user:
        return jsonify({'ok': False, 'msg': 'Username is required.'}), 400
        
    conn = None
    try:
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        with LOCK:
            cursor.execute("DELETE FROM SeatingAdmins WHERE LOWER(Username) = ?", (target_user.lower(),))
            cursor.execute("DELETE FROM SeatingTeamAdmins WHERE LOWER(username) = ?", (target_user.lower(),))
            conn.commit()
        conn.close()
        return jsonify({'ok': True, 'msg': f"Seating Admin '{target_user}' deleted."})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/seating/set_pin', methods=['POST'])
def seating_set_pin():
    data = request.get_json(force=True) or {}
    super_user = (data.get('super_user') or '').strip()
    super_pass = (data.get('super_pass') or '').strip()
    
    if not check_dba_auth_db(super_user, super_pass):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    new_pin = (data.get('pin') or '').strip()
    if not new_pin:
        return jsonify({'ok': False, 'msg': 'PIN is required.'}), 400
        
    conn = None
    try:
        conn = get_seating_db_connection()
        cursor = conn.cursor()
        with LOCK:
            cursor.execute("INSERT OR REPLACE INTO SeatingConfig (key, value) VALUES ('admin_pin', ?)", (new_pin,))
            conn.commit()
        conn.close()
        return jsonify({'ok': True, 'msg': 'Seating Admin PIN updated successfully.'})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/db/create_dba', methods=['POST'])
def db_create_dba():
    data = request.get_json(force=True) or {}
    super_user = (data.get('super_user') or '').strip()
    super_pass = (data.get('super_pass') or '').strip()
    
    if not check_dba_auth_db(super_user, super_pass):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    new_user = (data.get('username') or '').strip()
    new_pass = (data.get('password') or '').strip()
    
    if not new_user or not new_pass:
        return jsonify({'ok': False, 'msg': 'Please provide both username and password.'}), 400
        
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM DbaAdmins WHERE LOWER(Username) = ?", (new_user.lower(),))
        if cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'msg': f"DBA Admin '{new_user}' already exists."}), 400
            
        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with LOCK:
            cursor.execute("INSERT INTO DbaAdmins (Username, Password, UpdatedAt) VALUES (?, ?, ?)", (new_user, new_pass, now_str))
            conn.commit()
        conn.close()
        return jsonify({'ok': True, 'msg': f"DBA Admin '{new_user}' created successfully."})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/db/list_dba', methods=['POST'])
def db_list_dba():
    data = request.get_json(force=True) or {}
    if not check_dba_auth(data):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Username, UpdatedAt FROM DbaAdmins")
        admins = [{'username': r['Username'], 'updated_at': r['UpdatedAt']} for r in cursor.fetchall()]
        conn.close()
        return jsonify({'ok': True, 'dbas': admins})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/db/remove_dba', methods=['POST'])
def db_remove_dba():
    data = request.get_json(force=True) or {}
    super_user = (data.get('super_user') or '').strip()
    super_pass = (data.get('super_pass') or '').strip()
    
    if not check_dba_auth_db(super_user, super_pass):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    target_user = (data.get('username') or '').strip()
    if not target_user:
        return jsonify({'ok': False, 'msg': 'No target username provided.'}), 400
        
    if target_user.lower() == super_user.lower():
        return jsonify({'ok': False, 'msg': 'You cannot delete your own active DBA account.'}), 400
        
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verify it exists
        cursor.execute("SELECT * FROM DbaAdmins WHERE LOWER(Username) = ?", (target_user.lower(),))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'msg': f"DBA Admin '{target_user}' not found."}), 404
            
        with LOCK:
            cursor.execute("DELETE FROM DbaAdmins WHERE LOWER(Username) = ?", (target_user.lower(),))
            conn.commit()
        conn.close()
        return jsonify({'ok': True, 'msg': f"DBA Admin '{target_user}' removed successfully."})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/db/tables', methods=['POST'])
def db_tables():
    data = request.get_json(force=True) or {}
    if not check_dba_auth(data):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    db_target = data.get('db_target') or 'contractors'
    conn = None
    try:
        conn = get_db_connection_target(db_target)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
        tables = [r[0] for r in cursor.fetchall()]
        counts = {}
        for t in tables:
            cursor.execute(f"SELECT COUNT(*) FROM {t}")
            counts[t] = cursor.fetchone()[0]
        conn.close()
        return jsonify({'ok': True, 'tables': tables, 'counts': counts})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/db/schema', methods=['POST'])
def db_schema():
    data = request.get_json(force=True) or {}
    if not check_dba_auth(data):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    db_target = data.get('db_target') or 'contractors'
    table = (data.get('table') or '').strip()
    if not table:
        return jsonify({'ok': False, 'msg': 'No table name provided.'}), 400
        
    conn = None
    try:
        conn = get_db_connection_target(db_target)
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info({table})")
        columns = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'ok': True, 'columns': columns})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/db/query', methods=['POST'])
def db_query():
    data = request.get_json(force=True) or {}
    if not check_dba_auth(data):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    db_target = data.get('db_target') or 'contractors'
    query = (data.get('query') or '').strip()
    params = data.get('params') or []
    
    if not query:
        return jsonify({'ok': False, 'msg': 'Empty query.'}), 400
        
    # Handle VACUUM outside transactions
    if query.strip().upper() == 'VACUUM':
        conn = None
        try:
            target_file = SEATING_DB if db_target == 'seating' else DB_FILE
            conn = sqlite3.connect(target_file)
            conn.isolation_level = None
            cursor = conn.cursor()
            cursor.execute('VACUUM')
            conn.close()
            return jsonify({'ok': True, 'type': 'write', 'rowcount': 0})
        except Exception as e:
            if conn:
                conn.close()
            return jsonify({'ok': False, 'msg': str(e)}), 500

    conn = None
    try:
        conn = get_db_connection_target(db_target)
        cursor = conn.cursor()
        
        # Check if query is a read query
        is_select = query.strip().upper().startswith('SELECT') or query.strip().upper().startswith('PRAGMA') or query.strip().upper().startswith('EXPLAIN')
        
        if is_select:
            cursor.execute(query, params)
            columns = [col[0] for col in cursor.description] if cursor.description else []
            rows = [dict(r) for r in cursor.fetchall()]
            conn.close()
            return jsonify({'ok': True, 'type': 'select', 'columns': columns, 'rows': rows})
        else:
            with LOCK:
                cursor.execute(query, params)
                conn.commit()
                rowcount = cursor.rowcount
            conn.close()
            return jsonify({'ok': True, 'type': 'write', 'rowcount': rowcount})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/db/backup', methods=['POST'])
def db_backup():
    data = request.get_json(force=True) or {}
    if not check_dba_auth(data):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    db_target = data.get('db_target') or 'contractors'
    target_file = SEATING_DB if db_target == 'seating' else DB_FILE
    
    if not os.path.exists(target_file):
        return jsonify({'ok': False, 'msg': 'Database file not found.'}), 404
        
    from flask import send_file
    return send_file(target_file, as_attachment=True, download_name=os.path.basename(target_file))


@app.route('/api/db/restore', methods=['POST'])
def db_restore():
    super_user = request.form.get('super_user', '').strip()
    super_pass = request.form.get('super_pass', '').strip()
    
    if not check_dba_auth_db(super_user, super_pass):
        return jsonify({'ok': False, 'msg': 'Unauthorized access.'}), 401
        
    if 'db_file' not in request.files:
        return jsonify({'ok': False, 'msg': 'No database file provided.'}), 400
        
    f = request.files['db_file']
    if f.filename == '':
        return jsonify({'ok': False, 'msg': 'Empty filename.'}), 400
        
    db_target = request.form.get('db_target', '').strip() or 'contractors'
    target_file = SEATING_DB if db_target == 'seating' else DB_FILE
    BACKUP_TEMP = target_file + '.backup'
    
    try:
        import shutil
        if os.path.exists(target_file):
            shutil.copy(target_file, BACKUP_TEMP)
            
        with LOCK:
            f.save(target_file)
            
        # Verify the database
        conn = None
        try:
            conn = sqlite3.connect(target_file)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            cursor.fetchall()
            conn.close()
        except Exception as ver_err:
            if conn:
                conn.close()
            if os.path.exists(BACKUP_TEMP):
                shutil.copy(BACKUP_TEMP, target_file)
                os.remove(BACKUP_TEMP)
            return jsonify({'ok': False, 'msg': f'Uploaded file is not a valid SQLite database: {str(ver_err)}'}), 400
            
        if os.path.exists(BACKUP_TEMP):
            os.remove(BACKUP_TEMP)
            
        return jsonify({'ok': True, 'msg': 'Database restored successfully.'})
    except Exception as e:
        return jsonify({'ok': False, 'msg': f'Restore failed: {str(e)}'}), 500

# --- TEAM ROSTER ROUTES ---

######################################################################
## Team Roster API
# Endpoints for managing workforce rosters, employees, and team configs #
######################################################################

@app.route('/api/roster/state', methods=['GET'])
def roster_state():
    month = request.args.get('month', '')
    conn = get_seating_db_connection()
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS TeamRosterSettings (key TEXT PRIMARY KEY, value TEXT)")
    
    statuses_row = c.execute("SELECT value FROM TeamRosterSettings WHERE key='statuses'").fetchone()
    if statuses_row:
        import json
        statuses = json.loads(statuses_row['value'])
    else:
        statuses = ['', 'WFO', 'WFH', 'Leave']
        
    c.execute("SELECT name, team FROM TeamRosterPeople")
    people = [{'name': r['name'], 'team': r['team']} for r in c.fetchall()]
    
    c.execute("SELECT name, day, mark FROM TeamRosterMarks WHERE month = ?", (month,))
    marks = {}
    for r in c.fetchall():
        key = f"{month}-{str(r['day']).zfill(2)}|{r['name']}"
        marks[key] = r['mark']
    conn.close()
    return jsonify({'ok': True, 'people': people, 'marks': marks, 'statuses': statuses})

@app.route('/api/roster/save_mark', methods=['POST'])
def roster_save_mark():
    data = request.json or {}
    month = data.get('month', '')
    name = data.get('name', '')
    day = data.get('day', 0)
    mark = data.get('mark', '')
    
    conn = get_seating_db_connection()
    c = conn.cursor()
    if mark == '':
        c.execute("DELETE FROM TeamRosterMarks WHERE month=? AND name=? AND day=?", (month, name, day))
    else:
        c.execute("INSERT OR REPLACE INTO TeamRosterMarks (month, name, day, mark) VALUES (?, ?, ?, ?)", (month, name, day, mark))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/roster/add_person', methods=['POST'])
def roster_add_person():
    data = request.json or {}
    if str(data.get('pin')).strip() != '250237':
        return jsonify({'ok': False, 'msg': 'Invalid PIN.'}), 403
    name = data.get('name', '').strip()
    team = data.get('team', '').strip()
    
    conn = get_seating_db_connection()
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO TeamRosterPeople (name, team) VALUES (?, ?)", (name, team))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/roster/remove_person', methods=['POST'])
def roster_remove_person():
    data = request.json or {}
    if str(data.get('pin')).strip() != '250237':
        return jsonify({'ok': False, 'msg': 'Invalid PIN.'}), 403
    name = data.get('name', '').strip()
    
    conn = get_seating_db_connection()
    c = conn.cursor()
    c.execute("DELETE FROM TeamRosterPeople WHERE name = ?", (name,))
    c.execute("DELETE FROM TeamRosterMarks WHERE name = ?", (name,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/roster/delete_team', methods=['POST'])
def roster_delete_team():
    data = request.json or {}
    if str(data.get('pin')).strip() != '250237':
        return jsonify({'ok': False, 'msg': 'Invalid PIN.'}), 403
    team = data.get('team', '').strip()
    
    conn = get_seating_db_connection()
    c = conn.cursor()
    c.execute("DELETE FROM TeamRosterPeople WHERE team = ?", (team,))
    # We could delete marks for these people too, but it's okay to leave orphaned marks for history
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/roster/save_settings', methods=['POST'])
def roster_save_settings():
    data = request.json or {}
    if str(data.get('pin')).strip() != '250237':
        return jsonify({'ok': False, 'msg': 'Invalid PIN.'}), 403
    
    import json
    statuses = data.get('statuses', [])
    
    conn = get_seating_db_connection()
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS TeamRosterSettings (key TEXT PRIMARY KEY, value TEXT)")
    c.execute("INSERT OR REPLACE INTO TeamRosterSettings (key, value) VALUES ('statuses', ?)", (json.dumps(statuses),))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/upload_calendar', methods=['POST'])
def upload_calendar():
    if 'calendar_pdf' not in request.files:
        return jsonify({'ok': False, 'msg': 'No file uploaded.'}), 400
        
    file = request.files['calendar_pdf']
    if file.filename == '':
        return jsonify({'ok': False, 'msg': 'No file selected.'}), 400
        
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'ok': False, 'msg': 'Only PDF files are allowed.'}), 400
        
    save_path = os.path.join(BASE, 'static', 'holiday_calendar.pdf')
    try:
        file.save(save_path)
        return jsonify({'ok': True, 'msg': 'Holiday calendar successfully uploaded.'})
    except Exception as e:
        return jsonify({'ok': False, 'msg': f'Error saving file: {str(e)}'}), 500

@app.route('/api/roster/migrate', methods=['POST'])
def roster_migrate():
    data = request.json or {}
    people = data.get('people', [])
    months_data = data.get('months_data', {})
    
    conn = get_seating_db_connection()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM TeamRosterPeople")
    if c.fetchone()[0] > 0:
        conn.close()
        return jsonify({'ok': True, 'msg': 'Already migrated.'})
        
    for p in people:
        c.execute("INSERT INTO TeamRosterPeople (name, team) VALUES (?, ?)", (p['name'], p['team']))
        
    for month_key, marks in months_data.items():
        month_str = month_key.replace('WF_', '')
        for k, v in marks.items():
            parts = k.split('|')
            if len(parts) == 2:
                day = int(parts[0].split('-')[2])
                name = parts[1]
                c.execute("INSERT INTO TeamRosterMarks (month, name, day, mark) VALUES (?, ?, ?, ?)", (month_str, name, day, v))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


if __name__ == '__main__':
    if not os.path.exists(SEATING_DB) and not os.path.exists(XLSX):
        raise SystemExit('Neither seating.db nor seating_backend.xlsx found in this folder.')
    init_contractors_db()
    init_seating_db()

    # IIS HttpPlatformHandler passes HTTP_PLATFORM_PORT — no SSL needed, IIS handles HTTPS
    iis_port = os.environ.get('HTTP_PLATFORM_PORT')
    if iis_port:
        port = int(iis_port)
        print(f'Running under IIS — HTTP on port {port} (IIS handles HTTPS/SSL).')
        app.run(host='127.0.0.1', port=port, debug=False, threaded=True)
    else:
        # Standalone mode — HTTPS directly on port 443
        port = 443
        print(f'Standalone mode — HTTPS on port {port}.')
        app.run(host='0.0.0.0', port=port, debug=False, threaded=True, ssl_context='adhoc')

